"""
bom_parser.py

Parses Altium-exported BOM Excel (.xlsx) files into a list of BomLine
objects.

Design notes:
- Columns are read by their HEADER NAME, not by fixed position. Different
  BOM exports can have different column layouts (we've seen one export
  with an extra "Value" column that another export doesn't have), so
  relying on column position would silently read the wrong data.
- Only "Manufacturer Part Number" and "Quantity" are read. Every other
  column (Comment, Description, Designator, Value, LCSC codes, free-text
  notes, ...) is ignored -- we only need MPN to match against InvenTree's
  IPN field, and Quantity for the BomItem quantity.
- The project name is NOT read from inside the file -- it is the BOM
  file's name (without extension). See project_name_from_file().
"""

from dataclasses import dataclass
from pathlib import Path
from typing import List

import openpyxl


# Columns this tool absolutely needs to do its job. If either is missing
# from a BOM file, we refuse to guess and raise instead.
REQUIRED_COLUMNS = [
    "Manufacturer Part Number",
    "Quantity",
]


class BomParseError(Exception):
    """Raised when a BOM file is missing required columns or has bad data
    in a row (e.g. missing MPN or quantity)."""


@dataclass
class BomLine:
    """One row of a parsed BOM, i.e. one unique component used on the
    board, with the total quantity used across all its designators."""

    mpn: str        # Manufacturer Part Number -> matched against InvenTree IPN
    quantity: int


def project_name_from_file(file_path: str) -> str:
    """The project name is simply the BOM file's name, without its
    extension (e.g. 'PBA_SAR6_CB_MAIN_P0_B0_01304.xlsx' ->
    'PBA_SAR6_CB_MAIN_P0_B0_01304')."""
    return Path(file_path).stem


def parse_bom_file(file_path: str) -> List[BomLine]:
    """Read a single Altium BOM export and return one BomLine per row.

    Raises BomParseError if required columns are missing, or if a row is
    missing data we cannot proceed without (MPN or Quantity).
    """
    path = Path(file_path)
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active

    # Read only the header row first, so a BOM missing required columns
    # fails fast -- without ever reading the (potentially large) data
    # rows into memory.
    header_rows = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    if not header_rows:
        raise BomParseError(f"'{path.name}' is empty.")
    header = header_rows[0]

    missing_columns = [c for c in REQUIRED_COLUMNS if c not in header]
    if missing_columns:
        raise BomParseError(
            f"'{path.name}' is missing required column(s): "
            f"{', '.join(missing_columns)}"
        )

    mpn_col = header.index("Manufacturer Part Number")
    quantity_col = header.index("Quantity")

    lines: List[BomLine] = []
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        # Skip fully blank rows (can appear at the end of an export).
        if all(cell is None for cell in row):
            continue

        mpn = row[mpn_col]
        quantity = row[quantity_col]

        if not mpn or quantity is None:
            raise BomParseError(
                f"'{path.name}' row {row_number}: missing Manufacturer "
                f"Part Number or Quantity -- please fix the BOM file "
                f"before importing."
            )

        lines.append(BomLine(mpn=str(mpn).strip(), quantity=int(quantity)))

    return lines