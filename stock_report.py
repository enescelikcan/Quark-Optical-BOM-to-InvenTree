"""
stock_report.py

Writes a stock-availability report for a BOM to a new .xlsx file: one
row per BOM line, with each part's current InvenTree stock ("In Stock")
next to it.

This module only knows how to WRITE the report -- it doesn't read or
match anything itself. The caller (StockReportWorker in main.py)
resolves each BOM line's Comment against InvenTree and hands this
module a plain list of (comment, stock) pairs to write out. Keeping
this separate mirrors bom_parser.py (reads BOM files) and
inventree_client.py (talks to InvenTree): each module does exactly one
thing.
"""

from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font

NOT_FOUND_LABEL = "Not found"


def write_stock_report(output_path: str, rows: List[Tuple[str, Optional[float]]]) -> None:
    """Write a stock report to `output_path`.

    `rows` is a list of (comment, stock) pairs, in the order they
    should appear. `stock` is the part's current InvenTree "In Stock"
    quantity, or None if the Comment could not be matched to any
    InvenTree part -- written as "Not found" rather than left blank,
    so it's unambiguous when opened later.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Stock Report"

    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")

    sheet["A1"] = "Comment"
    sheet["B1"] = "In Stock"
    sheet["A1"].font = header_font
    sheet["B1"].font = header_font

    for row_index, (comment, stock) in enumerate(rows, start=2):
        sheet.cell(row=row_index, column=1, value=comment).font = body_font
        value = stock if stock is not None else NOT_FOUND_LABEL
        sheet.cell(row=row_index, column=2, value=value).font = body_font

    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 14

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
